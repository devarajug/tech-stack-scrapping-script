## required imports
import os
import sys
import json
import warnings
import requests
import pandas as pd
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook
from requests.auth import HTTPProxyAuth
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

class TechStack:

    def __init__(self):
        warnings.filterwarnings('ignore')
        self.url = "https://nvd.nist.gov/vuln/search/results?query="
        self.noOfIssuesCount = None
        self.countFrom = None
        self.countThrough = None
        self.startIndex = 0
        self.today = datetime.today().strftime('%d-%m-%Y')
        self.base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.auditor_comments_file = "" #filename if auditor comments file avaliable
        self.output_file_name = 'dependency-check-report-Jile-' + self.today + '.xlsx'
        self.cpeMatchStrings = {
            "spring_framework_vmware 4.3.25" : "cpe:2.3:a:vmware:spring_framework:4.3.25",
            "spring_framework_pivotal 4.3.25" : "cpe:/a:pivotal_software:spring_framework:4.3.25",
            "PostgreSQL 11.11" : "cpe:/a:postgresql:postgresql:11.11",
            "Amazon Corretto 1.8.0_252" : "cpe:/a:oracle:jdk:1.8.0:update_252",
            "Tomcat 9.0.58" : "cpe:/a:apache:tomcat:9.0.28"
        }

    def getDataFromWeb(self, url):
        try:
            request = requests.Session()
            request.verify = False
            data = request.get(url=url)

        except Exception as e:
            sys.exit(e)
        return data

    def readAuditorCommentsFile(self):
        if os.path.isfile(self.auditor_comments_file):
            try:
                with open(self.auditor_comments_file, 'r') as rb:
                    comments_data = json.loads(rb.read())
            except Exception as e:
                sys.exit(e)
        else:
            comments_data = {}
        return comments_data

    def scrapeTechStackData(self, cpe, startIndex=0):
        try:
            url = self.url + cpe + "&startIndex=" + str(startIndex)
            data = self.getDataFromWeb(url=url)
            parsed_data = BeautifulSoup(data.text, 'lxml')
            self.noOfIssuesCount = int(parsed_data.select_one('strong[data-testid=vuln-matching-records-count]').text)
            self.countFrom = int(parsed_data.select_one('strong[data-testid=vuln-displaying-count-from]').text)
            self.countThrough = int(parsed_data.select_one('strong[data-testid=vuln-displaying-count-through]').text)
            TechStackData = parsed_data.select_one('table[data-testid=vuln-results-table]')
        except Exception as e:
            sys.exit(e)
        return TechStackData

    def techStackDataToDf(self, comments_data):
        try:
            print()
            print("Analysis Started. It Takes Time to Complete, Please Wait Patiently")
            productname, cve, severity, description, auditor_comment, status = [[] for i in range(6)]
            for product, cpe in self.cpeMatchStrings.items():
                print()
                data = self.scrapeTechStackData(cpe=cpe)
                if self.noOfIssuesCount == 0:
                    productname.append(product.strip())
                    cve.append("No vulnerability")
                    severity.append("No vulnerability")
                    description.append("No vulnerability")
                    auditor_comment.append('No vulnerability')
                    status.append("Closed")
                    print(productname[-1]+ " : " + cve[-1] + " : " + severity[-1])
                elif self.noOfIssuesCount <= 20:
                    issues_table = self.scrapeTechStackData(cpe=cpe)
                    non_dispute_issues_count=0
                    for i in range(self.noOfIssuesCount):
                        description_data = issues_table.select_one('tr[data-testid=vuln-row-'+str(i)+'] td p[data-testid=vuln-summary-'+str(i)+']').text.strip()
                        if "unspecified vulnerability" in description_data.lower() or "disputed" in description_data.lower():
                            continue
                        else:
                            productname.append(product.strip())
                            cve.append(issues_table.select_one('tr[data-testid=vuln-row-'+str(i)+'] th strong a[href]').text.strip())
                            description.append(description_data)
                            cvss3 = issues_table.select_one('tr[data-testid=vuln-row-'+str(i)+'] td[nowrap=nowrap] span[id=cvss3-link]')
                            if cvss3:
                                cvss3_score_severity = cvss3.text.split(":")[-1]
                                cvss3_severity = cvss3_score_severity.split(" ")[-1]
                                severity.append(cvss3_severity.strip())
                            else:
                                cvss2 = issues_table.select_one('tr[data-testid=vuln-row-'+str(i)+'] td[nowrap=nowrap] span[id=cvss2-link]').text
                                cvss2_score_severity = cvss2.split(":")[-1]
                                cvss2_severity = cvss2_score_severity.split(" ")[-1]
                                severity.append(cvss2_severity.strip())
                            status.append(comments_data.get(productname[-1], {}).get(cve[-1], {}).get("Status", "Open"))
                            auditor_comment.append(comments_data.get(productname[-1], {}).get(cve[-1], {}).get("Comment", ""))
                            non_dispute_issues_count+=1
                        print(productname[-1]+ " : " + cve[-1] + " : " + severity[-1])
                    else:
                        if non_dispute_issues_count == 0:
                            productname.append(product)
                            cve.append("No vulnerability")
                            severity.append("No vulnerability")
                            description.append("No vulnerability")
                            auditor_comment.append('No vulnerability')
                            status.append("Closed")
                            print(productname[-1]+ " : " + cve[-1] + " : " + severity[-1])
                elif self.noOfIssuesCount > 20:
                    count_while = 0
                    while self.noOfIssuesCount - self.startIndex >= 0 :
                        issues_table = self.scrapeTechStackData(cpe=cpe, startIndex=self.startIndex)
                        for i in range(self.countThrough+1 - self.countFrom):
                            description_data = issues_table.select_one('tr[data-testid=vuln-row-'+str(i)+'] td p[data-testid=vuln-summary-'+str(i)+']').text.strip()
                            if "unspecified vulnerability" in description_data.lower() or "disputed" in description_data.lower():
                                continue
                            else:
                                productname.append(product.strip())
                                cve.append(issues_table.select_one('tr[data-testid=vuln-row-'+str(i)+'] th strong a[href]').text.strip())
                                description.append(description_data)
                                cvss3 = issues_table.select_one('tr[data-testid=vuln-row-'+str(i)+'] td[nowrap=nowrap] span[id=cvss3-link]')
                                if cvss3:
                                    cvss3_score_severity = cvss3.text.split(":")[-1]
                                    cvss3_severity = cvss3_score_severity.split(" ")[-1]
                                    severity.append(cvss3_severity.strip())
                                else:
                                    cvss2 = issues_table.select_one('tr[data-testid=vuln-row-'+str(i)+'] td[nowrap=nowrap] span[id=cvss2-link]').text
                                    cvss2_score_severity = cvss2.split(":")[-1]
                                    cvss2_severity = cvss2_score_severity.split(" ")[-1]
                                    severity.append(cvss2_severity.strip())
                                status.append(comments_data.get(productname[-1], {}).get(cve[-1], {}).get("Status", "Open"))
                                auditor_comment.append(comments_data.get(productname[-1], {}).get(cve[-1], {}).get("Comment", ""))
                                count_while+=1
                            print(productname[-1]+ " : " + cve[-1] + " : " + severity[-1])
                        self.startIndex+=20
                    else:
                        if count_while == 0:
                            productname.append(product)
                            cve.append("No vulnerability")
                            severity.append("No vulnerability")
                            description.append("No vulnerability")
                            auditor_comment.append('No vulnerability')
                            status.append("Closed")
                            print(productname[-1]+ " : " + cve[-1] + " : " + severity[-1])
                else:
                    sys.exit("some thing went wrong pls re run the script")
            result_data_tech_stack = zip(productname,description,cve,severity, status, auditor_comment)
            df_tech_stack = pd.DataFrame(
                list(result_data_tech_stack),
                columns = ['Product','Description','CVE','Severity', 'Status', 'Auditor Comment']
            )
        except Exception as e:
            df_tech_stack = None
            sys.exit(e)

        return df_tech_stack

    def makeXL(self, df_tech_stack):
        try:
            workbook = Workbook()
            workbook.remove(workbook.active)
            header_font = Font(name='Calibri',bold=True,color='FFFFFF')
            centered_alignment = Alignment(horizontal='center')
            wrapped_alignment = Alignment(vertical='top',wrap_text=False)
            fill = PatternFill(start_color='5FABE6',end_color='5FABE6',fill_type='solid',)

            tech_stack_sheet_columns = [
                ('Product', 40),
                ('Description', 40),
                ('CVE', 30),
                ('Severity', 15),
                ('Status', 15),
                ('Auditor Comment', 40),
                ('Developer Comment', 40)
            ]

            worksheet = workbook.create_sheet(title='TechStack',index=0)
            row_num = 1
            for col_num, (column_title, column_width) in enumerate(tech_stack_sheet_columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                cell.alignment = centered_alignment
                cell.fill = fill
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                column_dimensions.width = column_width
            for i in range(len(df_tech_stack)):
                row_num += 1
                row = [
                    (df_tech_stack.loc[i,'Product'],'Normal'),
                    (df_tech_stack.loc[i,'Description'],'Normal'),
                    (df_tech_stack.loc[i,'CVE'],'Normal'),
                    (df_tech_stack.loc[i,'Severity'],'Normal'),
                    (df_tech_stack.loc[i,'Status'],'Normal'),
                    (df_tech_stack.loc[i,'Auditor Comment'],'Normal')
                ]
                for col_num, (cell_value, cell_format) in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.style = cell_format
                    cell.alignment = wrapped_alignment

            worksheet.freeze_panes = worksheet['A2']
            worksheet.sheet_properties.tabColor = '5FABE6'
            workbook.save(self.output_file_name)
            print()
            print('execl created successfully....')
            print()
            return
        except Exception as e:
            print("Unable to create xls....")
            sys.exit(e)

if __name__ == "__main__":

    df_tech_stack, comments_data = None, {}
    TPC = TechStack()
    df_tech_stack = TPC.techStackDataToDf(comments_data=comments_data)
    TPC.makeXL(df_tech_stack=df_tech_stack)
